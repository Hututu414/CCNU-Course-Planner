from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import Course
from .time_parser import parse_meetings
from .utils import clean_text, normalize_header, parse_float


FIELD_ALIASES: dict[str, list[str]] = {
    "name": ["课程名称", "课程名", "名称"],
    "course_id": ["课程编号", "课程代码", "课程号", "编号"],
    "teacher": ["授课教师", "任课教师", "主讲教师", "教师"],
    "credit": ["课程学分", "学分"],
    "raw_time": ["上课时间", "上课时间地点", "时间安排", "上课节次"],
    "classroom": ["上课地点", "教室", "地点"],
    "campus": ["上课校区", "校区"],
    "exam_type": ["考核方式", "考试方式"],
}

CATEGORY_HEADERS = ["课程性质", "开课类型", "开课单位", "授课对象", "备注"]
HEADER_SCAN_ROWS = 100
BLANK_ROW_STOP_LIMIT = 80


def load_courses_from_excel(path: str) -> list[Course]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    courses: list[Course] = []
    try:
        for worksheet in workbook.worksheets:
            header_row, headers = _find_header_row(worksheet)
            if header_row is None:
                continue
            courses.extend(_read_sheet_courses(worksheet, header_row, headers))
    finally:
        workbook.close()
    return courses


def _find_header_row(worksheet) -> tuple[int | None, dict[str, int]]:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True),
        start=1,
    ):
        headers = {
            normalize_header(value): column_index
            for column_index, value in enumerate(row)
            if normalize_header(value)
        }
        if _header_has(headers, "课程名称") and _header_has(headers, "上课时间"):
            return row_index, headers
    return None, {}


def _read_sheet_courses(worksheet, header_row: int, headers: dict[str, int]) -> list[Course]:
    field_columns = {
        field_name: _find_column(headers, aliases)
        for field_name, aliases in FIELD_ALIASES.items()
    }
    category_columns = [
        column
        for header in CATEGORY_HEADERS
        if (column := _find_column(headers, [header])) is not None
    ]

    courses: list[Course] = []
    blank_streak = 0
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        values = [clean_text(value) for value in row]
        if not any(values):
            blank_streak += 1
            if blank_streak >= BLANK_ROW_STOP_LIMIT:
                break
            continue
        blank_streak = 0

        name = _value(values, field_columns.get("name"))
        if not name or name == "课程名称" or name.startswith("说明"):
            continue

        raw_time = _value(values, field_columns.get("raw_time"))
        meetings = parse_meetings(raw_time)
        parse_warning = ""
        if raw_time and not meetings:
            parse_warning = "无法解析上课时间"
        elif not raw_time:
            parse_warning = "缺少上课时间"

        category = " / ".join(
            value for value in (_value(values, column) for column in category_columns) if value
        )

        courses.append(
            Course(
                source_sheet=worksheet.title,
                source_row=row_index,
                course_id=_value(values, field_columns.get("course_id")),
                name=name,
                teacher=_value(values, field_columns.get("teacher")),
                credit=parse_float(_value(values, field_columns.get("credit"))),
                raw_time=raw_time,
                campus=_value(values, field_columns.get("campus")),
                classroom=_value(values, field_columns.get("classroom")),
                exam_type=_value(values, field_columns.get("exam_type")),
                category=category,
                meetings=meetings,
                parse_warning=parse_warning,
            )
        )
    return courses


def _header_has(headers: dict[str, int], expected: str) -> bool:
    return _find_column(headers, [expected]) is not None


def _find_column(headers: dict[str, int], aliases: list[str]) -> int | None:
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        if alias in headers:
            return headers[alias]
    for header, column in headers.items():
        if any(alias in header or header in alias for alias in normalized_aliases):
            return column
    return None


def _value(values: list[str], column: int | None) -> str:
    if column is None or column >= len(values):
        return ""
    return values[column]
